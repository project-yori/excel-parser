from openpyxl import load_workbook
import json
import os

# Get paths of files to parse
path = './filesToParse/member'
filesToParse = []

for files in os.walk(path):
  for file in files[2]:
    if('.xlsx' in file):
      filesToParse.append('{}/{}'.format(path, file))

if(not os.path.exists('./output')):
  os.mkdir('output')

def parse(filesToParse):
  for file in filesToParse:
    wb = load_workbook(file)
    ws = wb[wb.sheetnames[0]]
        
    members = []

    # Read properties
    props = []
    for column in range(1, 8):
      props.append(ws.cell(row=1, column=column).value)

    # Count number of rows
    numOfRow = 1
    while(ws.cell(row=numOfRow, column=1).value != None):
      numOfRow += 1

    # Iterate cells and save into array
    for row in range(2, numOfRow):
      newMember = {}
      for column in range(0, 7):
        data = ws.cell(row=row, column=column+1).value
        if(props[column]=='graduated'):
          data = True if data == 1 else False
        newMember[props[column]] = data
      members.append(newMember)

    # Write to json
    fileName = file.replace('.xlsx', '').replace(path+'/', '');
    with open('output/{}.json'.format(fileName), 'w') as output:
      json.dump(members, output, ensure_ascii=False);
    
parse(filesToParse)
