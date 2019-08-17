from openpyxl import load_workbook
import json
import os

# Get paths of files to parse
path = './filesToParse/costume'
filesToParse = []

for files in os.walk(path):
  for file in files[2]:
    if('.xlsx' in file):
      filesToParse.append('{}/{}'.format(path, file))

if(not os.path.exists('./output')):
  os.mkdir('output')

def parse(filesToParse):
  for file in filesToParse:
    wb = load_workbook(file)
    ws = wb[wb.sheetnames[0]]
        
    costumes = []

    # Read properties
    props = []
    for column in range(1, 4):
      props.append(ws.cell(row=1, column=column).value)

    # Count number of rows
    numOfRow = 1
    while(ws.cell(row=numOfRow, column=1).value != None):
      numOfRow += 1

    # Iterate cells and save into array
    for row in range(2, numOfRow):
      newCos = {}
      for column in range(0, 3):
        data = ws.cell(row=row, column=column+1).value
        newCos[props[column]] = data
      costumes.append(newCos)

    # Write to json
    fileName = file.replace('.xlsx', '').replace(path+'/', '');
    with open('output/{}.json'.format(fileName), 'w') as output:
      json.dump(costumes, output, ensure_ascii=False);
    
parse(filesToParse)
