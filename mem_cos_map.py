from openpyxl import load_workbook
import json
import os

# Get paths of files to parse
path = './filesToParse/mem_cos_map'
filesToParse = []

for files in os.walk(path):
  for file in files[2]:
    if('.xlsx' in file):
      filesToParse.append('{}/{}'.format(path, file))

if(not os.path.exists('./output')):
  os.mkdir('output')


def parse(filesToParse):
  for file in filesToParse:
    wb = load_workbook(file)
    ws = wb[wb.sheetnames[0]]
        
  mem_cos_map = {}

  # Count number of members and save each key
  members = []
  numOfMem = 1
  while(ws.cell(row=2, column=numOfMem+1).value != None):
    members.append(ws.cell(row=2, column=numOfMem+1).value.lower())
    numOfMem += 1
  
  # Count number of costumes and save each key
  costumes = []
  numOfCos = 1
  while(ws.cell(row=numOfCos+2, column=1).value != None):
    costumes.append(ws.cell(row=numOfCos+2, column=1).value)
    numOfCos += 1
    
  # Iterate cells and save into array
  for column in range(2, numOfMem):
    new_mem_cos_obj = []
    for row in range(3, numOfCos):
      if(ws.cell(row=row, column=column).value == 1):
        new_mem_cos_obj.append(costumes[row-3])
    mem_cos_map[members[column-2]] = new_mem_cos_obj

  # Write to json
  fileName = file.replace('.xlsx', '').replace(path+'/', '');
  with open('output/{}.json'.format(fileName), 'w') as output:
    json.dump(mem_cos_map, output, ensure_ascii=False)   


parse(filesToParse)
